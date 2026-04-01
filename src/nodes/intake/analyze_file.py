import os
import subprocess
import sys


FILE_TYPE_DEPS = {
    ".csv": [],
    ".tsv": [],
    ".json": [],
    ".parquet": ["pyarrow"],
    ".xlsx": ["openpyxl"],
    ".xls": ["xlrd"],
    ".feather": ["pyarrow"],
}

SUPPORTED_TYPES = list(FILE_TYPE_DEPS.keys())


def analyze_file(state: dict) -> dict:
    """Analyze file size, type, and estimated memory. Install deps if needed."""
    filepath = state["filepath"]

    if not os.path.exists(filepath):
        raise FileNotFoundError(f"File not found: {filepath}")

    file_size = os.path.getsize(filepath)
    file_ext = os.path.splitext(filepath)[1].lower()

    if file_ext not in FILE_TYPE_DEPS:
        raise ValueError(f"Unsupported file type: {file_ext}. Supported: {SUPPORTED_TYPES}")

    # Install any missing dependencies for this file type
    missing_deps = _check_deps(FILE_TYPE_DEPS[file_ext])
    if missing_deps:
        _install_deps(missing_deps)

    # Peek at file for row/column estimate
    row_count, columns = _peek_file(filepath, file_ext)

    # Pandas typically uses 2-5x file size in memory
    memory_estimate_mb = round((file_size * 3) / (1024 * 1024), 2)

    state["nodes"]["analyze_file"] = {
        "filepath": filepath,
        "file_ext": file_ext,
        "file_size_bytes": file_size,
        "file_size_mb": round(file_size / (1024 * 1024), 2),
        "row_count": row_count,
        "column_count": len(columns),
        "column_names": columns,
        "estimated_memory_mb": memory_estimate_mb,
        "installed_deps": missing_deps,
    }

    print(f"   File: {filepath} ({file_ext})")
    print(f"   Size: {state['nodes']['analyze_file']['file_size_mb']} MB")
    print(f"   Rows: ~{row_count}, Columns: {len(columns)}")
    print(f"   Estimated memory: ~{memory_estimate_mb} MB")
    if missing_deps:
        print(f"   Installed: {', '.join(missing_deps)}")
    return state


def _check_deps(deps: list) -> list:
    """Return list of deps that aren't installed."""
    missing = []
    for dep in deps:
        try:
            __import__(dep)
        except ImportError:
            missing.append(dep)
    return missing


def _install_deps(deps: list) -> None:
    """Install missing packages into the current environment."""
    print(f"   Installing dependencies: {', '.join(deps)}")
    subprocess.run([sys.executable, "-m", "pip", "install"] + deps + ["-q"], check=True)


def _peek_file(filepath: str, file_ext: str) -> tuple:
    """Get row count and column names without loading the full file."""
    if file_ext in (".csv", ".tsv"):
        sep = "," if file_ext == ".csv" else "\t"
        row_count = 0
        columns = []
        with open(filepath, "r") as f:
            header = f.readline().strip()
            columns = header.split(sep)
            for _ in f:
                row_count += 1
        return row_count, columns

    if file_ext == ".json":
        import json
        with open(filepath, "r") as f:
            data = json.load(f)
        if isinstance(data, list):
            columns = list(data[0].keys()) if data else []
            return len(data), columns
        return 1, list(data.keys())

    # For binary formats, do a minimal load to get schema
    if file_ext == ".parquet":
        import pyarrow.parquet as pq
        pf = pq.ParquetFile(filepath)
        columns = pf.schema.names
        return pf.metadata.num_rows, columns

    if file_ext in (".xlsx", ".xls"):
        import openpyxl
        wb = openpyxl.load_workbook(filepath, read_only=True)
        ws = wb.active
        columns = [cell.value for cell in next(ws.iter_rows(max_row=1))]
        row_count = ws.max_row - 1
        wb.close()
        return row_count, columns

    if file_ext == ".feather":
        import pyarrow.feather as feather
        import pyarrow as pa
        table = feather.read_table(filepath)
        return table.num_rows, table.column_names

    return 0, []
